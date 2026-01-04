"""
インポート処理用のユーティリティモジュール
CSV/Excelファイルの読み込み、データ変換、出納帳への投入を行う
"""

import csv
import json
from datetime import datetime
from io import StringIO, BytesIO
from openpyxl import load_workbook
from db import SessionLocal
from models import CashBook, AccountItem


class ImportProcessor:
    """インポート処理を管理するクラス"""
    
    def __init__(self):
        self.errors = []
        self.warnings = []
        self.imported_count = 0
    
    def read_csv_file(self, file_content, encoding='utf-8'):
        """CSVファイルを読み込む"""
        try:
            # ファイル内容を文字列に変換
            if isinstance(file_content, bytes):
                try:
                    text = file_content.decode(encoding)
                except UnicodeDecodeError:
                    # Shift-JISで試す
                    text = file_content.decode('shift_jis')
            else:
                text = file_content
            
            # CSVを読み込む
            reader = csv.reader(StringIO(text))
            rows = list(reader)
            return rows
        except Exception as e:
            self.errors.append(f"CSV読み込みエラー: {str(e)}")
            return None
    
    def read_excel_file(self, file_content):
        """Excelファイルを読み込む"""
        try:
            # BytesIOオブジェクトに変換
            if isinstance(file_content, bytes):
                file_obj = BytesIO(file_content)
            else:
                file_obj = file_content
            
            # Excelを読み込む
            wb = load_workbook(file_obj)
            ws = wb.active
            
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            
            return rows
        except Exception as e:
            self.errors.append(f"Excel読み込みエラー: {str(e)}")
            return None
    
    def get_preview_data(self, file_content, file_type, skip_rows=0, limit=5):
        """ファイルのプレビューデータを取得（最初の数行）"""
        try:
            if file_type == 'csv':
                rows = self.read_csv_file(file_content)
            elif file_type == 'excel':
                rows = self.read_excel_file(file_content)
            else:
                self.errors.append("サポートされていないファイル形式です")
                return None
            
            if rows is None:
                return None
            
            # スキップ行を除外
            preview_rows = rows[skip_rows:skip_rows + limit]
            return preview_rows
        except Exception as e:
            self.errors.append(f"プレビュー取得エラー: {str(e)}")
            return None
    
    def parse_date(self, date_str):
        """日付文字列をパース"""
        if not date_str:
            return None
        
        # 複数の日付フォーマットを試す
        date_formats = [
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%Y年%m月%d日',
            '%m-%d-%Y',
            '%m/%d/%Y',
            '%d-%m-%Y',
            '%d/%m/%Y',
        ]
        
        for fmt in date_formats:
            try:
                dt = datetime.strptime(str(date_str).strip(), fmt)
                return dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
        
        self.warnings.append(f"日付形式が不正です: {date_str}")
        return None
    
    def parse_amount(self, amount_str):
        """金額文字列をパース"""
        if not amount_str:
            return 0
        
        try:
            # カンマを削除
            amount_str = str(amount_str).replace(',', '').strip()
            # 括弧で囲まれた負数に対応
            if amount_str.startswith('(') and amount_str.endswith(')'):
                return -int(float(amount_str[1:-1]))
            return int(float(amount_str))
        except ValueError:
            self.warnings.append(f"金額形式が不正です: {amount_str}")
            return 0
    
    def import_data(self, file_content, file_type, mapping, skip_rows=0, account_item_id=None):
        """
        ファイルから出納帳データをインポート
        
        Args:
            file_content: ファイル内容（bytes or str）
            file_type: ファイル形式（'csv' or 'excel'）
            mapping: マッピング情報（辞書）
                {
                    'date_col': 0,           # 取引日の列番号
                    'amount_col': 1,         # 金額の列番号
                    'counterparty_col': 2,   # 取引先の列番号（任意）
                    'remarks_col': 3,        # 摘要の列番号（任意）
                    'account_item_id': 5     # 勘定科目ID（固定値）
                }
            skip_rows: スキップするヘッダー行数
            account_item_id: 勘定科目ID（マッピングで指定されない場合）
        
        Returns:
            dict: インポート結果
        """
        db = SessionLocal()
        self.errors = []
        self.warnings = []
        self.imported_count = 0
        
        try:
            # ファイルを読み込む
            if file_type == 'csv':
                rows = self.read_csv_file(file_content)
            elif file_type == 'excel':
                rows = self.read_excel_file(file_content)
            else:
                self.errors.append("サポートされていないファイル形式です")
                return self._get_result()
            
            if rows is None:
                return self._get_result()
            
            # スキップ行を除外
            data_rows = rows[skip_rows:]
            
            # マッピング情報を取得
            date_col = mapping.get('date_col')
            amount_col = mapping.get('amount_col')
            counterparty_col = mapping.get('counterparty_col')
            remarks_col = mapping.get('remarks_col')
            mapped_account_id = mapping.get('account_item_id')
            
            # 必須マッピングの確認
            if date_col is None or amount_col is None:
                self.errors.append("取引日と金額のマッピングは必須です")
                return self._get_result()
            
            # 勘定科目の決定
            if mapped_account_id:
                final_account_id = mapped_account_id
            elif account_item_id:
                final_account_id = account_item_id
            else:
                self.errors.append("勘定科目が指定されていません")
                return self._get_result()
            
            # 勘定科目の存在確認
            account = db.query(AccountItem).filter(
                AccountItem.id == final_account_id
            ).first()
            if not account:
                self.errors.append(f"勘定科目ID {final_account_id} が見つかりません")
                return self._get_result()
            
            # データを投入
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            for row_idx, row in enumerate(data_rows, start=skip_rows + 1):
                try:
                    # 列が足りない場合はスキップ
                    if len(row) <= max(date_col, amount_col):
                        self.warnings.append(f"行 {row_idx}: 列数が不足しています")
                        continue
                    
                    # 取引日を取得
                    transaction_date = self.parse_date(row[date_col])
                    if not transaction_date:
                        self.warnings.append(f"行 {row_idx}: 取引日が無効です")
                        continue
                    
                    # 金額を取得
                    amount = self.parse_amount(row[amount_col])
                    if amount == 0:
                        self.warnings.append(f"行 {row_idx}: 金額が無効です")
                        continue
                    
                    # オプション項目を取得
                    counterparty = None
                    if counterparty_col is not None and len(row) > counterparty_col:
                        counterparty = str(row[counterparty_col]).strip() if row[counterparty_col] else None
                    
                    remarks = None
                    if remarks_col is not None and len(row) > remarks_col:
                        remarks = str(row[remarks_col]).strip() if row[remarks_col] else None
                    
                    # 重複チェック（同じ日付・金額・取引先の組み合わせ）
                    existing = db.query(CashBook).filter(
                        CashBook.transaction_date == transaction_date,
                        CashBook.amount_with_tax == amount,
                        CashBook.account_item_id == final_account_id,
                        CashBook.counterparty == counterparty
                    ).first()
                    
                    if existing:
                        self.warnings.append(
                            f"行 {row_idx}: 同じ日付・金額・取引先のデータが既に存在します"
                        )
                        continue
                    
                    # 出納帳に投入
                    cash_book = CashBook(
                        transaction_date=transaction_date,
                        account_item_id=final_account_id,
                        counterparty=counterparty,
                        remarks=remarks,
                        amount_with_tax=amount,
                        created_at=now,
                        updated_at=now
                    )
                    db.add(cash_book)
                    self.imported_count += 1
                
                except Exception as e:
                    self.errors.append(f"行 {row_idx}: {str(e)}")
                    continue
            
            # コミット
            db.commit()
            
        except Exception as e:
            db.rollback()
            self.errors.append(f"インポート処理エラー: {str(e)}")
        finally:
            db.close()
        
        return self._get_result()
    
    def _get_result(self):
        """結果を辞書形式で返す"""
        return {
            'success': len(self.errors) == 0,
            'imported_count': self.imported_count,
            'errors': self.errors,
            'warnings': self.warnings
        }
